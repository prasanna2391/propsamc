# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.decorators import api_view, authentication_classes, permission_classes, parser_classes
from rest_framework.permissions import AllowAny
# Create your views here.
import googlemaps
import openpyxl
from rest_framework import status

@api_view(['POST'])
@parser_classes((FormParser, MultiPartParser,))
@authentication_classes([AllowAny])
@permission_classes([AllowAny])
def excel(request):
    file_obj = request.FILES['file']
    wb = openpyxl.load_workbook(file_obj)
    sheet = wb.get_sheet_names()[0]
    datasheet = wb.get_sheet_by_name(sheet)
    gmaps = googlemaps.Client(key='AIzaSyBd8hk_wjI8MS6NmNUGKgaCkxDCsyxdof8')
    for i in range(1,datasheet.max_row +1):
       try:
           address = datasheet.cell(row=i,column =1).value
           geocode_result = gmaps.geocode(address)
           latitude, longitude = geocode_result[0]['geometry']['location']['lat'], \
                                 geocode_result[0]['geometry']['location']['lng']
           datasheet.cell(row=i, column=2).value = latitude
           datasheet.cell(row=i, column=3).value = longitude
       except:
           continue
    wb.save('../media_cdn/' + file_obj.name)
    filepath = '../media_cdn/' + file_obj.name
    return Response(filepath)